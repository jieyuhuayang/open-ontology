"""File import service — upload preview and background import for Excel/CSV."""

import asyncio
import csv
import io
import logging
import os
import time
import uuid
from datetime import datetime, timezone

from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.domain.common import generate_rid
from app.domain.constants import DEFAULT_ONTOLOGY_RID, DEFAULT_USER_ID
from app.domain.import_task import ImportTask, ImportTaskStatus
from app.domain.type_mapping import infer_column_type
from app.exceptions import AppError
from app.services.import_task_service import shared_import_task_service as _import_task_service

logger = logging.getLogger(__name__)

_ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}
_PREVIEW_TTL_SECONDS = 30 * 60  # 30 minutes

# Module-level preview cache: {token: (data_dict, created_at)}
_previews: dict[str, tuple[dict, datetime]] = {}


def _cleanup_expired_previews() -> None:
    """Remove preview entries older than TTL."""
    now = datetime.now(timezone.utc)
    expired = [
        token
        for token, (_, created_at) in _previews.items()
        if (now - created_at).total_seconds() > _PREVIEW_TTL_SECONDS
    ]
    for token in expired:
        _previews.pop(token, None)


class FileImportService:
    def __init__(self, session: AsyncSession):
        self._session = session

    async def upload_and_preview(
        self,
        filename: str,
        file_content: bytes,
        content_type: str,
    ) -> dict:
        _cleanup_expired_previews()

        # Validate file size
        max_bytes = settings.UPLOAD_MAX_SIZE_MB * 1024 * 1024
        if len(file_content) > max_bytes:
            raise AppError(
                code="FILE_TOO_LARGE",
                message=f"File exceeds maximum size of {settings.UPLOAD_MAX_SIZE_MB}MB",
                status_code=400,
            )

        # Validate extension
        ext = os.path.splitext(filename)[1].lower()
        if ext not in _ALLOWED_EXTENSIONS:
            raise AppError(
                code="UNSUPPORTED_FILE_FORMAT",
                message=f"Unsupported file format: {ext}. Supported: .csv, .xlsx, .xls",
                status_code=400,
            )

        # Save to temp directory
        file_token = uuid.uuid4().hex
        temp_dir = os.path.join(settings.UPLOAD_TEMP_DIR, file_token)
        os.makedirs(temp_dir, exist_ok=True)
        temp_path = os.path.join(temp_dir, filename)
        with open(temp_path, "wb") as f:
            f.write(file_content)

        # Parse preview
        if ext == ".csv":
            preview = self._parse_csv_preview(file_content)
            sheets = None
            default_sheet = None
        elif ext == ".xlsx":
            preview, sheets, default_sheet = self._parse_xlsx_preview(temp_path)
        elif ext == ".xls":
            preview, sheets, default_sheet = self._parse_xls_preview(temp_path)
        else:
            raise AppError(
                code="UNSUPPORTED_FILE_FORMAT",
                message=f"Unsupported: {ext}",
                status_code=400,
            )

        # Store preview data for confirm step (with creation timestamp for TTL)
        _previews[file_token] = (
            {
                "path": temp_path,
                "ext": ext,
                "filename": filename,
                "preview": preview,
            },
            datetime.now(timezone.utc),
        )

        return {
            "fileToken": file_token,
            "filename": filename,
            "fileSize": len(file_content),
            "sheets": sheets,
            "defaultSheet": default_sheet,
            "preview": preview,
        }

    def _parse_csv_preview(self, content: bytes) -> dict:
        text = content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        all_rows = list(reader)

        if not all_rows:
            return {"columns": [], "rows": [], "totalRows": 0, "hasHeader": True}

        header = all_rows[0]
        data_rows = all_rows[1:]

        # Infer column types from data values
        columns = []
        for i, col_name in enumerate(header):
            values = [row[i] if i < len(row) else "" for row in data_rows[:1000]]
            inferred = infer_column_type(values)
            columns.append(
                {
                    "name": col_name,
                    "inferredType": inferred,
                    "sampleValues": values[:3],
                }
            )

        # Build preview rows as dicts
        preview_rows = []
        for row in data_rows[:50]:
            row_dict = {}
            for i, col in enumerate(header):
                row_dict[col] = row[i] if i < len(row) else ""
            preview_rows.append(row_dict)

        return {
            "columns": columns,
            "rows": preview_rows,
            "totalRows": len(data_rows),
            "hasHeader": True,
        }

    def _parse_xlsx_preview(self, path: str) -> tuple[dict, list[str], str]:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        default_sheet = sheets[0] if sheets else ""
        ws = wb[default_sheet]

        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])
        wb.close()

        if not rows:
            return (
                {"columns": [], "rows": [], "totalRows": 0, "hasHeader": True},
                sheets,
                default_sheet,
            )

        header = rows[0]
        data_rows = rows[1:]

        columns = []
        for i, col_name in enumerate(header):
            values = [r[i] if i < len(r) else "" for r in data_rows[:1000]]
            inferred = infer_column_type(values)
            columns.append(
                {
                    "name": col_name,
                    "inferredType": inferred,
                    "sampleValues": values[:3],
                }
            )

        preview_rows = []
        for row in data_rows[:50]:
            row_dict = {header[i]: row[i] if i < len(row) else "" for i in range(len(header))}
            preview_rows.append(row_dict)

        preview = {
            "columns": columns,
            "rows": preview_rows,
            "totalRows": len(data_rows),
            "hasHeader": True,
        }
        return preview, sheets, default_sheet

    def _parse_xls_preview(self, path: str) -> tuple[dict, list[str], str]:
        import python_calamine

        wb = python_calamine.CalamineWorkbook.from_path(path)
        sheets = wb.sheet_names
        default_sheet = sheets[0] if sheets else ""

        ws = wb.get_sheet_by_name(default_sheet)
        rows_data = ws.to_python()

        if not rows_data:
            return (
                {"columns": [], "rows": [], "totalRows": 0, "hasHeader": True},
                sheets,
                default_sheet,
            )

        header = [str(cell) if cell is not None else "" for cell in rows_data[0]]
        data_rows = [
            [str(cell) if cell is not None else "" for cell in row] for row in rows_data[1:]
        ]

        columns = []
        for i, col_name in enumerate(header):
            values = [r[i] if i < len(r) else "" for r in data_rows[:1000]]
            inferred = infer_column_type(values)
            columns.append(
                {
                    "name": col_name,
                    "inferredType": inferred,
                    "sampleValues": values[:3],
                }
            )

        preview_rows = []
        for row in data_rows[:50]:
            row_dict = {header[i]: row[i] if i < len(row) else "" for i in range(len(header))}
            preview_rows.append(row_dict)

        preview = {
            "columns": columns,
            "rows": preview_rows,
            "totalRows": len(data_rows),
            "hasHeader": True,
        }
        return preview, sheets, default_sheet

    async def confirm_import(
        self,
        file_token: str,
        dataset_name: str,
        sheet_name: str | None = None,
        has_header: bool = True,
        selected_columns: list[str] | None = None,
        column_type_overrides: dict[str, str] | None = None,
    ) -> ImportTask:
        entry = _previews.get(file_token)
        if not entry:
            raise AppError(
                code="FILE_TOKEN_EXPIRED",
                message="File token not found or expired",
                status_code=400,
            )
        preview_data = entry[0]

        task = _import_task_service.create_task()

        asyncio.create_task(
            self._run_import(
                task.task_id,
                preview_data,
                dataset_name,
                sheet_name,
                has_header,
                selected_columns,
                column_type_overrides,
            )
        )
        return task

    async def _run_import(
        self,
        task_id: str,
        preview_data: dict,
        dataset_name: str,
        sheet_name: str | None,
        has_header: bool,
        selected_columns: list[str] | None,
        column_type_overrides: dict[str, str] | None,
    ) -> None:
        start_time = time.monotonic()
        _import_task_service.update_status(task_id, ImportTaskStatus.RUNNING)

        try:
            from app.database import async_session_factory
            from app.storage.dataset_storage import DatasetStorage

            path = preview_data["path"]
            ext = preview_data["ext"]

            # Re-read file
            if ext == ".csv":
                header, data_rows = self._read_csv_full(path, has_header)
            elif ext == ".xlsx":
                header, data_rows = self._read_xlsx_full(path, sheet_name, has_header)
            elif ext == ".xls":
                header, data_rows = self._read_xls_full(path, sheet_name, has_header)
            else:
                raise ValueError(f"Unsupported extension: {ext}")

            # Filter columns
            if selected_columns:
                col_indices = [i for i, h in enumerate(header) if h in selected_columns]
                header = [header[i] for i in col_indices]
                data_rows = [
                    [row[i] if i < len(row) else "" for i in col_indices] for row in data_rows
                ]

            # Build column definitions
            columns_info = []
            for i, col_name in enumerate(header):
                values = [r[i] if i < len(r) else "" for r in data_rows[:1000]]
                inferred = infer_column_type(values)
                if column_type_overrides and col_name in column_type_overrides:
                    inferred = column_type_overrides[col_name]
                columns_info.append(
                    {
                        "name": col_name,
                        "inferred_type": inferred,
                    }
                )

            # Build rows
            rows_data = []
            for row in data_rows:
                row_dict = {header[i]: row[i] if i < len(row) else "" for i in range(len(header))}
                rows_data.append(row_dict)

            # Save to database
            dataset_rid = generate_rid("ontology", "dataset")
            source_metadata = {
                "sourceFilename": preview_data["filename"],
                "hasHeader": has_header,
            }
            if sheet_name:
                source_metadata["sheetName"] = sheet_name

            source_type = "csv" if ext == ".csv" else "excel"

            async with async_session_factory() as session:
                async with session.begin():
                    await DatasetStorage.create(
                        session,
                        dataset_rid=dataset_rid,
                        name=dataset_name,
                        source_type=source_type,
                        source_metadata=source_metadata,
                        ontology_rid=DEFAULT_ONTOLOGY_RID,
                        created_by=DEFAULT_USER_ID,
                        columns=columns_info,
                        rows=rows_data,
                    )

            duration = int((time.monotonic() - start_time) * 1000)
            _import_task_service.update_status(
                task_id,
                ImportTaskStatus.COMPLETED,
                dataset_rid=dataset_rid,
                row_count=len(rows_data),
                column_count=len(columns_info),
                duration_ms=duration,
            )

            # Cleanup temp file
            try:
                os.unlink(path)
                os.rmdir(os.path.dirname(path))
            except OSError:
                pass

        except Exception as e:
            duration = int((time.monotonic() - start_time) * 1000)
            _import_task_service.update_status(
                task_id,
                ImportTaskStatus.FAILED,
                error_code="FILE_IMPORT_FAILED",
                error_message=str(e),
                duration_ms=duration,
            )

    def _read_csv_full(self, path: str, has_header: bool) -> tuple[list[str], list[list[str]]]:
        with open(path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            all_rows = list(reader)
        if not all_rows:
            return [], []
        if has_header:
            return all_rows[0], all_rows[1:]
        else:
            header = [f"column_{i}" for i in range(len(all_rows[0]))]
            return header, all_rows

    def _read_xlsx_full(
        self, path: str, sheet_name: str | None, has_header: bool
    ) -> tuple[list[str], list[list[str]]]:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])
        wb.close()
        if not rows:
            return [], []
        if has_header:
            return rows[0], rows[1:]
        else:
            header = [f"column_{i}" for i in range(len(rows[0]))]
            return header, rows

    def _read_xls_full(
        self, path: str, sheet_name: str | None, has_header: bool
    ) -> tuple[list[str], list[list[str]]]:
        import python_calamine

        wb = python_calamine.CalamineWorkbook.from_path(path)
        ws = wb.get_sheet_by_name(sheet_name or wb.sheet_names[0])
        rows_data = ws.to_python()
        rows = [[str(cell) if cell is not None else "" for cell in row] for row in rows_data]
        if not rows:
            return [], []
        if has_header:
            return rows[0], rows[1:]
        else:
            header = [f"column_{i}" for i in range(len(rows[0]))]
            return header, rows
